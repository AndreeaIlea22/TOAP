import csv

from openpyxl import load_workbook
from portality.models import Journal

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("-o", "--out", help="output file path")
    args = parser.parse_args()

    if not args.out:
        print("Please specify an output file path with the -o option")
        parser.print_help()
        exit()

    with open(args.out, "w", encoding="utf-8") as f:
        writer = csv.writer(f)

        file = './resources/SAGE_jnls_in_DOAJ.edited.xlsx'
        wb = load_workbook(file)
        sheet = wb['sage_journals']

        for r in range(1, sheet.max_row+1):
            j = Journal.pull(sheet.cell(row=r, column=1).value)
            if j is not None:
                if sheet.cell(row=r, column=1).value != j.id or sheet.cell(row=r, column=2).value != j.bibjson().title:
                    #if title of the journal in the sheet and in the system do not match - ignore
                    writer.writerow(["Id of requested journal does not match its title. Id: " +
                                    sheet.cell(row=r, column=1).value + ", journal ignored"])
                else:
                    fulltext_url = sheet.cell(row=r, column=3).value
                    apc_url = sheet.cell(row=r, column=4).value
                    submission_url = sheet.cell(row=r, column=5).value
                    editorial_board_url = sheet.cell(row=r, column=6).value
                    review_process_url = sheet.cell(row=r, column=7).value
                    aims_scope_url = sheet.cell(row=r, column=8).value
                    author_instructions = sheet.cell(row=r, column=9).value
                    plagiarism_url = sheet.cell(row=r, column=10).value
                    oa_url = sheet.cell(row=r, column=11).value
                    license_url = sheet.cell(row=r, column=12).value

                    jbib = j.bibjson()

                    if fulltext_url is not None:
                        jbib.remove_urls("homepage")
                        jbib.add_url(fulltext_url, "homepage")

                    if apc_url is not None:
                        jbib.apc_url = apc_url

                    if submission_url is not None:
                        jbib.submission_charges_url = submission_url

                    if editorial_board_url is not None:
                        jbib.remove_urls("editorial_board")
                        jbib.add_url(editorial_board_url, "editorial_board")

                    if review_process_url is not None:
                        jbib.set_editorial_review(jbib.editorial_review["process"], editorial_board_url)

                    if aims_scope_url is not None:
                        jbib.remove_urls("aims_scope")
                        jbib.add_url(aims_scope_url, "aims_scope")

                    if author_instructions is not None:
                        jbib.remove_urls("author_instructions")
                        jbib.add_url(author_instructions, "author_instructions")

                    if plagiarism_url is not None:
                        jbib.set_plagiarism_detection(plagiarism_url)

                    if oa_url is not None:
                        jbib.remove_urls("oa_statement")
                        jbib.add_url(oa_url, "oa_statement")

                    if license_url is not None:
                        current_license = jbib.get_license()
                        jbib.set_license(license_title=current_license["title"],
                                         license_type=current_license["type"],
                                         url=license_url,
                                         open_access=current_license["open_access"],
                                         by=current_license["BY"],
                                         sa=current_license["SA"],
                                         nc=current_license["NC"],
                                         nd=current_license["ND"],
                                         embedded=current_license["embedded"],
                                         embedded_example_url=current_license["embedded_example_url"])
                    j.save(blocking=True)

            else:
                # if journal's id is not found in the system
                writer.writerow(["Journal not found: " + sheet.cell(row=r, column=1).value])

    # finished
    writer.writerow(["Finished."])
